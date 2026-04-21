"""Smoke tests for scoring/compile_xlsx.py (US-003)."""
import json
from pathlib import Path

import openpyxl
import pytest

from scoring import compile_xlsx


def _synth_extraction() -> dict:
    return {
        "deal": {
            "TargetName": "Acme Corp",
            "Acquirer": "Buyer LLC",
            "DateAnnounced": "2020-04-01",
            "DateEffective": "2020-06-30",
            "DateFiled": "2020-04-05",
            "FormType": "DEFM14A",
            "URL": "https://example.com/filing",
            "auction": True,
            "bidder_registry": {
                "bidder_01": {"resolved_name": "Alpha Financial", "aliases_observed": ["Party A"]},
                "bidder_02": {"resolved_name": "Beta Strategic Inc.", "aliases_observed": ["Party B"]},
            },
        },
        "events": [
            {
                "BidderID": 1,
                "bidder_name": "bidder_01",
                "bidder_alias": "Party A",
                "bidder_type": {"base": "f", "non_us": False, "public": False},
                "bid_note": "NDA",
                "bid_date_precise": "2020-01-10",
                "source_quote": "Party A signed a confidentiality agreement.",
                "source_page": 5,
                "comments": "approach; preliminary contact",
            },
            {
                "BidderID": 2,
                "bidder_name": "bidder_02",
                "bidder_alias": "Party B",
                "bidder_type": {"base": "s", "non_us": True, "public": True},
                "bid_note": "Executed",
                "bid_value_pershare": 42.50,
                "bid_value_unit": "USD_per_share",
                "bid_date_precise": "2020-04-01",
                "all_cash": True,
                "source_quote": "Party B and Acme executed the merger agreement.",
                "source_page": 42,
                "comments": "",
            },
        ],
    }


def test_flatten_bidder_type_financial():
    ev = {"bidder_type": {"base": "f", "non_us": False, "public": False}}
    flat = compile_xlsx.flatten_bidder_type(ev)
    assert flat["bidder_type_financial"] == 1
    assert flat["bidder_type_strategic"] == 0
    assert flat["bidder_type_mixed"] == 0
    assert flat["bidder_type_nonUS"] == 0
    assert flat["bidder_type_note"] == "private"


def test_flatten_bidder_type_strategic_public_nonus():
    ev = {"bidder_type": {"base": "s", "non_us": True, "public": True}}
    flat = compile_xlsx.flatten_bidder_type(ev)
    assert flat["bidder_type_strategic"] == 1
    assert flat["bidder_type_financial"] == 0
    assert flat["bidder_type_nonUS"] == 1
    assert flat["bidder_type_note"] == "public"


def test_flatten_bidder_type_missing_ok():
    flat = compile_xlsx.flatten_bidder_type({})
    assert flat["bidder_type_financial"] == 0
    assert flat["bidder_type_strategic"] == 0
    assert flat["bidder_type_mixed"] == 0
    assert flat["bidder_type_nonUS"] == 0
    assert flat["bidder_type_note"] == ""


def test_pristine_has_35_columns(tmp_path):
    out = tmp_path / "pristine.xlsx"
    rows = compile_xlsx.compile_to_xlsx_pristine([_synth_extraction()], str(out))
    assert rows == 2

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert len(header) == 35
    assert header[:5] == ["Row", "TargetName", "gvkeyT", "DealNumber", "Acquirer"]
    assert "source_quote" not in header
    assert "source_page" not in header


def test_audit_has_37_columns(tmp_path):
    out = tmp_path / "audit.xlsx"
    compile_xlsx.compile_to_xlsx_audit([_synth_extraction()], str(out))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert len(header) == 37
    assert header[-2] == "source_page"
    assert header[-1] == "source_quote"


def test_rejects_raw_json_missing_events(tmp_path):
    out = tmp_path / "bad.xlsx"
    with pytest.raises(ValueError, match="events"):
        compile_xlsx.compile_to_xlsx_pristine([{"deal": {}}], str(out))


def test_downstream_merge_columns_are_blank(tmp_path):
    out = tmp_path / "merge.xlsx"
    compile_xlsx.compile_to_xlsx_pristine([_synth_extraction()], str(out))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    gvkeyT_col = header.index("gvkeyT") + 1
    gvkeyA_col = header.index("gvkeyA") + 1
    dealnum_col = header.index("DealNumber") + 1
    cshoc_col = header.index("cshoc") + 1
    for row in range(2, 4):
        assert ws.cell(row=row, column=gvkeyT_col).value in ("", None)
        assert ws.cell(row=row, column=gvkeyA_col).value in ("", None)
        assert ws.cell(row=row, column=dealnum_col).value in ("", None)
        assert ws.cell(row=row, column=cshoc_col).value in ("", None)


def test_comments_split_on_semicolons(tmp_path):
    out = tmp_path / "comments.xlsx"
    compile_xlsx.compile_to_xlsx_pristine([_synth_extraction()], str(out))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    c1_col = header.index("comments_1") + 1
    c2_col = header.index("comments_2") + 1
    assert ws.cell(row=2, column=c1_col).value == "approach"
    assert ws.cell(row=2, column=c2_col).value == "preliminary contact"
