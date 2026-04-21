"""Project validated extraction JSON into Alex's 35-column xlsx format.

Ported from bids_pipeline/pipeline/compile.py as part of US-003 of the
2026-04-21 validator-hardening PRD. Two adaptations from the source:

  1. bids_try uses a nested `bidder_type = {base, non_us, public}` object;
     Alex's layout uses 5 flat columns (`bidder_type_financial`,
     `bidder_type_strategic`, `bidder_type_mixed`, `bidder_type_nonUS`,
     `bidder_type_note`). `flatten_bidder_type()` does the projection.

  2. bids_try extractions carry `source_quote` and `source_page` per row;
     Alex's workbook does not. The `pristine` xlsx drops them to match
     Alex exactly; the `audit` xlsx appends them as columns 36-37 for
     reviewer spot-check. Per PRD §7.5 recommendation (c).

Downstream-merge columns (`gvkeyT`, `gvkeyA`, `DealNumber`, `cshoc`) are
left blank; fabricating values there would poison the merge.
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Alex's 35 columns in exact order.
COLUMNS_PRISTINE = [
    ("Row", 6),
    ("TargetName", 28),
    ("gvkeyT", 10),
    ("DealNumber", 12),
    ("Acquirer", 28),
    ("gvkeyA", 10),
    ("DateAnnounced", 14),
    ("DateEffective", 14),
    ("DateFiled", 14),
    ("FormType", 12),
    ("URL", 40),
    ("Auction", 8),
    ("BidderID", 10),
    ("BidderName", 30),
    ("bidder_type_financial", 10),
    ("bidder_type_strategic", 10),
    ("bidder_type_mixed", 10),
    ("bidder_type_nonUS", 10),
    ("bidder_type_note", 15),
    ("bid_value", 12),
    ("bid_value_pershare", 14),
    ("bid_value_lower", 14),
    ("bid_value_upper", 14),
    ("bid_value_unit", 12),
    ("multiplier", 10),
    ("bid_type", 10),
    ("bid_date_precise", 14),
    ("bid_date_rough", 14),
    ("bid_note", 22),
    ("all_cash", 8),
    ("additional_note", 20),
    ("cshoc", 8),
    ("comments_1", 40),
    ("comments_2", 40),
    ("comments_3", 40),
]

# Audit file: pristine + 2 citation columns.
COLUMNS_AUDIT = COLUMNS_PRISTINE + [
    ("source_page", 12),
    ("source_quote", 80),
]


def flatten_bidder_type(event: dict) -> dict:
    """Map nested `bidder_type = {base, non_us, public}` to Alex's 5 flat cols.

    Returns a dict with bidder_type_financial / _strategic / _mixed / _nonUS
    booleans (0/1) and bidder_type_note (string OR empty).
    """
    bt = event.get("bidder_type") or {}
    base = bt.get("base") if isinstance(bt, dict) else None
    non_us = bt.get("non_us") if isinstance(bt, dict) else None
    public = bt.get("public") if isinstance(bt, dict) else None

    financial = 1 if base == "f" else 0
    strategic = 1 if base == "s" else 0
    mixed = 1 if base == "mixed" else 0
    non_us_flag = 1 if non_us else 0

    # Alex's legacy note column is free text; we use it to surface the
    # `public` flag since Alex does not carry a separate public column.
    # Empty string when no info.
    note_parts = []
    if public is True:
        note_parts.append("public")
    elif public is False:
        note_parts.append("private")
    note = "; ".join(note_parts) if note_parts else ""

    return {
        "bidder_type_financial": financial,
        "bidder_type_strategic": strategic,
        "bidder_type_mixed": mixed,
        "bidder_type_nonUS": non_us_flag,
        "bidder_type_note": note,
    }


def _split_comments(event: dict) -> tuple[str, str, str]:
    """bids_try collapses comments_1/2/3 into one field; re-split on semicolons.

    If there are more than 3 segments, the tail is joined into comments_3.
    """
    raw = event.get("comments")
    if not isinstance(raw, str) or not raw.strip():
        return "", "", ""
    parts = [p.strip() for p in raw.split(";") if p.strip()]
    c1 = parts[0] if len(parts) >= 1 else ""
    c2 = parts[1] if len(parts) >= 2 else ""
    c3 = "; ".join(parts[2:]) if len(parts) >= 3 else ""
    return c1, c2, c3


def _stringify_list_field(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "; ".join(str(v) for v in value)
    return str(value)


def _require_ready(extraction: dict) -> None:
    """Fail closed when handed raw JSON (missing top-level shape)."""
    if not isinstance(extraction, dict):
        raise ValueError("extraction must be a dict")
    deal = extraction.get("deal")
    events = extraction.get("events")
    if not isinstance(deal, dict):
        raise ValueError("extraction['deal'] must be a dict")
    if not isinstance(events, list):
        raise ValueError("extraction['events'] must be a list")


def extraction_to_rows(extraction: dict, url: str = "", include_audit: bool = False) -> list[dict]:
    """Convert one extraction to row dicts."""
    _require_ready(extraction)
    deal = extraction.get("deal") or {}

    target = deal.get("TargetName") or deal.get("target") or ""
    acquirer = deal.get("Acquirer") or deal.get("acquirer") or ""
    date_announced = deal.get("DateAnnounced") or deal.get("date_announced") or ""
    date_effective = deal.get("DateEffective") or deal.get("date_effective") or ""
    date_filed = deal.get("DateFiled") or deal.get("date_filed") or ""
    form_type = deal.get("FormType") or ""
    auction = 1 if deal.get("auction") else 0
    deal_url = url or deal.get("URL") or ""

    rows: list[dict] = []
    for event in extraction.get("events") or []:
        c1, c2, c3 = _split_comments(event)
        bidder_type_flat = flatten_bidder_type(event)

        row: dict = {
            "TargetName": target,
            "gvkeyT": "",
            "DealNumber": "",
            "Acquirer": acquirer,
            "gvkeyA": "",
            "DateAnnounced": date_announced,
            "DateEffective": date_effective,
            "DateFiled": date_filed,
            "FormType": form_type,
            "URL": deal_url,
            "Auction": auction,
            "BidderID": event.get("BidderID"),
            "BidderName": event.get("bidder_name", ""),
            **bidder_type_flat,
            "bid_value": event.get("bid_value"),
            "bid_value_pershare": event.get("bid_value_pershare"),
            "bid_value_lower": event.get("bid_value_lower"),
            "bid_value_upper": event.get("bid_value_upper"),
            "bid_value_unit": event.get("bid_value_unit", "") or "",
            "multiplier": "",
            "bid_type": event.get("bid_type", "") or "",
            "bid_date_precise": event.get("bid_date_precise", "") or "",
            "bid_date_rough": event.get("bid_date_rough", "") or "",
            "bid_note": event.get("bid_note", "") or "",
            "all_cash": event.get("all_cash"),
            "additional_note": event.get("additional_note", "") or "",
            "cshoc": "",
            "comments_1": c1,
            "comments_2": c2,
            "comments_3": c3,
        }
        if include_audit:
            row["source_page"] = _stringify_list_field(event.get("source_page"))
            row["source_quote"] = _stringify_list_field(event.get("source_quote"))
        rows.append(row)
    return rows


def _write_workbook(extractions: list[dict], output_path: str, columns: list[tuple], include_audit: bool) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "deal_details"

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    row_num = 2
    global_row = 1
    deal_idx = 0

    for extraction in extractions:
        try:
            deal_rows = extraction_to_rows(extraction, include_audit=include_audit)
        except ValueError as exc:
            label = (extraction.get("deal") or {}).get("TargetName") or "?"
            raise ValueError(f"Deal {label}: {exc}") from None
        deal_idx += 1

        for dr in deal_rows:
            ws.cell(row=row_num, column=1, value=global_row)
            for col_idx, (col_name, _) in enumerate(columns, 1):
                if col_name == "Row":
                    continue
                val = dr.get(col_name)
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                if deal_idx % 2 == 0:
                    cell.fill = alt_fill
            global_row += 1
            row_num += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    return row_num - 2


def compile_to_xlsx_pristine(extractions: list[dict], output_path: str) -> int:
    """Alex's exact 35 columns, no citation fields."""
    return _write_workbook(extractions, output_path, COLUMNS_PRISTINE, include_audit=False)


def compile_to_xlsx_audit(extractions: list[dict], output_path: str) -> int:
    """Pristine + source_page + source_quote for reviewer spot-check."""
    return _write_workbook(extractions, output_path, COLUMNS_AUDIT, include_audit=True)


def compile_to_xlsx(extractions: list[dict], output_path: str) -> int:
    """Convenience alias = compile_to_xlsx_pristine (Alex-facing default)."""
    return compile_to_xlsx_pristine(extractions, output_path)


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print(
            "Usage: python scoring/compile_xlsx.py <output.xlsx> <extraction1.json> [extraction2.json ...]\n"
            "Environment: set SEC_EXTRACT_AUDIT=1 to emit a 37-column audit file instead of pristine."
        )
        sys.exit(1)

    output_path = sys.argv[1]
    extraction_files = sys.argv[2:]

    extractions = []
    for fpath in extraction_files:
        extractions.append(json.loads(Path(fpath).read_text()))

    import os
    audit = bool(os.environ.get("SEC_EXTRACT_AUDIT"))
    try:
        n = (compile_to_xlsx_audit if audit else compile_to_xlsx_pristine)(
            extractions, output_path
        )
    except ValueError as exc:
        print(f"X {exc}", file=sys.stderr)
        sys.exit(1)
    kind = "audit (37 cols)" if audit else "pristine (35 cols)"
    print(f"compiled {len(extractions)} deals, {n} rows -> {output_path} [{kind}]")
